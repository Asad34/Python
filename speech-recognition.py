import speech_recognition as sr
import sounddevice as sd
from scipy.io.wavfile import write
import wavio as wv
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# -----------------------------------------------------------------------
# Sampling frequency
freq = 44100

# Recording duration
duration = 5

# Start recorder with the given values
# of duration and sample frequency
recording = sd.rec(int(duration * freq),
                   samplerate=freq, channels=2)

# Record audio for the given number of seconds
sd.wait()

# Convert the NumPy array to audio file
wv.write("recording1.wav", recording, freq, sampwidth=2)

filename = "recording1.wav"
# initialize the recognizer
r = sr.Recognizer()
# open the file
with sr.AudioFile(filename) as source:
    # listen for the data (load audio to memory)
    audio_data = r.record(source)
    # recognize (convert from speech to text)
    text = r.recognize_google(audio_data)
    print(text)

# string to search in file
word = text
with open(r'students.txt', 'r') as fp:
    # read all lines in a list
    lines = fp.readlines()
    for line in lines:
        # check if string present on a current line
        if line.find(word) != -1:
            print('Line Number:', lines.index(line))
            print('student:', line)
# ---------------------------------------------------------------------------------------

# create a workbook object
# wb = Workbook()
# load an existing spreadsheet
wb = load_workbook('biology.xlsx')
# create an active worksheet
ws = wb.active
# search data
roll_numbers = ws['A']
tempRoll = text
index = 0
for roll in roll_numbers:
    index += 1
    if roll.value == int(tempRoll):
        break
student = ws[index]

for data in student:
    print(data.value, end=" ")
